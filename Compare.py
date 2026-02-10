# -*- coding: utf-8 -*-
# Quarter vs Quarter Highlighter GUI (OB Main ID key = column B)
# -------------------------------------------------------------
# - Pick First (older) and Second (newer) Excel workbooks
# - Sheet name defaults to first sheet
# - Match rows by OB Main ID from column B
# - Compare user-selected Excel column letters (required)
# - Highlight changes in Second workbook (light blue)
# - Highlight cells where Q1 has value and Q2 is blank (pink)
# - For columns O, S, T: if changed, highlight differing words in Q2 in red
# - Highlight entire row for new projects in Second workbook (light blue)
# - Output saved next to Second workbook as: <Q2 name> (change highlighted).xlsx
#
# Requirements:
#   pip install pandas openpyxl
#
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont  # <-- important for rich text fonts

APP_TITLE = "Quarter Change Highlighter"

# Columns where we want word-level red diff when values change
WORD_DIFF_COLUMNS = {"O", "S", "T"}

# ---- helpers ----
def normalize_colname(name: str) -> str:
    if name is None:
        return ""
    s = re.sub(r"\s+", " ", str(name)).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")

def read_df(path, sheet_name=None):
    # Read everything as strings to avoid dtype surprises
    return pd.read_excel(path, sheet_name=sheet_name, dtype=str, engine="openpyxl")

def to_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def header_from_letter(ws, letter: str) -> str:
    """Read the header text at the given Excel column letter from row 1."""
    idx = column_index_from_string(letter)
    return ws.cell(row=1, column=idx).value or ""

def build_lookup(df, key_header):
    lookup = {}
    if key_header not in df.columns:
        return lookup
    for _, row in df.iterrows():
        key = to_str(row.get(key_header, ""))
        if key:
            lookup[key] = row  # keep whole row for convenience
    return lookup

def get_headers_for_letters(ws, letters):
    """Map Excel column letters to header strings (assumes header row is row 1)."""
    header_map = {}
    for L in letters:
        idx = column_index_from_string(L)
        header_cell = ws.cell(row=1, column=idx).value
        header_map[L] = header_cell if header_cell is not None else ""
    return header_map

def find_matching_headers_in_q1(df_q1, q2_headers):
    """Given Second Workbook header names at letters, find the same in First Workbook (exact, else normalized)."""
    q1_cols = list(df_q1.columns)
    q1_norm_map = {normalize_colname(c): c for c in q1_cols}
    matched = {}
    for L, h in q2_headers.items():
        if h in q1_cols:
            matched[L] = h
        else:
            n = normalize_colname(h)
            matched[L] = q1_norm_map.get(n, h)  # fallback if missing
    return matched

def apply_word_diff(cell, old_text: str, new_text: str):
    """
    Replace cell value with rich text where words that differ from old_text
    are colored red, and unchanged words remain default color.
    Comparison is by word position (split on whitespace).
    """
    old_tokens = old_text.split()
    new_tokens = new_text.split()

    rt = CellRichText()
    for i, token in enumerate(new_tokens):
        # InlineFont is required for TextBlock.font
        if i < len(old_tokens) and token == old_tokens[i]:
            font = InlineFont()  # default
        else:
            font = InlineFont(color="FFFF0000")  # red

        text = token
        if i < len(new_tokens) - 1:
            text += " "

        rt.append(TextBlock(font=font, text=text))

    cell.value = rt

def highlight_q2_changes(
    q1_path,
    q2_path,
    target_letters,  # REQUIRED: tuple/list of column letters
    sheet_name=None,
    progress_cb=None,
):
    if not target_letters:
        raise ValueError("No columns provided to compare. Please enter column letters (e.g., I,O,P,Q,R).")

    # Load Second workbook and sheet
    if progress_cb: progress_cb("Loading Second workbook...")
    wb = load_workbook(q2_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    sheet_name = ws.title

    # Key header from column B (OB Main ID)
    key_header_q2 = header_from_letter(ws, "B")
    if not key_header_q2:
        raise RuntimeError(
            "Column B header is empty in the Second workbook. "
            "Please ensure column B contains the OB Main ID header."
        )

    # Read as DataFrames
    if progress_cb: progress_cb("Reading sheets...")
    df_q2 = read_df(q2_path, sheet_name=sheet_name)
    df_q1 = read_df(q1_path, sheet_name=sheet_name)

    # Find corresponding key header in First workbook
    if key_header_q2 in df_q1.columns:
        key_header_q1 = key_header_q2
    else:
        q1_norm_map = {normalize_colname(c): c for c in df_q1.columns}
        key_header_q1 = q1_norm_map.get(normalize_colname(key_header_q2))
        if key_header_q1 is None:
            raise RuntimeError(
                f"Could not find the OB Main ID header from column B ('{key_header_q2}') in the First workbook. "
                "Ensure both workbooks share the same OB Main ID header."
            )

    # Build lookups
    if progress_cb: progress_cb("Indexing by OB Main ID...")
    q1_lookup = build_lookup(df_q1, key_header_q1)

    # Map target letters to headers in Q2, and find matching in Q1
    q2_headers_by_letter = get_headers_for_letters(ws, target_letters)
    q1_headers_by_letter = find_matching_headers_in_q1(df_q1, q2_headers_by_letter)

    # Styles (light blue & pink; ARGB)
    light_blue = "FF94DCF8"
    light_pink = "FFED8EDA"
    changed_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
    cleared_fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")  # Q1 value, Q2 blank
    newrow_fill  = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")

    if progress_cb: progress_cb("Scanning rows...")
    max_row = ws.max_row
    max_col = ws.max_column

    # Iterate rows in Q2 (skip header row 1)
    for r in range(2, max_row + 1):
        df_idx = r - 2
        if not (0 <= df_idx < len(df_q2)):
            continue

        key_val = to_str(df_q2.iloc[df_idx].get(key_header_q2, ""))
        is_new_project = key_val and (key_val not in q1_lookup)

        if is_new_project:
            # Highlight entire row for new projects
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = newrow_fill
            continue

        # If exists in Q1, compare target columns
        if key_val in q1_lookup:
            q1_row = q1_lookup[key_val]
            for L in target_letters:
                col_idx = column_index_from_string(L)
                cell2 = ws.cell(row=r, column=col_idx)
                q2_val = to_str(cell2.value)
                q2_header = q2_headers_by_letter.get(L, None)
                q1_header = q1_headers_by_letter.get(L, q2_header)

                q1_val = ""
                if q1_header in df_q1.columns:
                    q1_val = to_str(q1_row.get(q1_header, ""))

                # Case 1: Q1 has value, Q2 is blank -> pink
                if q1_val != "" and q2_val == "":
                    cell2.fill = cleared_fill
                # Case 2: Values differ (normal change) -> blue
                elif q2_val != q1_val:
                    cell2.fill = changed_fill

                    # Extra: word-level diff for O, S, T
                    if L in WORD_DIFF_COLUMNS and q2_val != "":
                        apply_word_diff(cell2, q1_val, q2_val)

    # Save output beside Q2 as "<Q2 name> (change highlighted).xlsx"
    q2_dir = os.path.dirname(q2_path)
    q2_base = os.path.basename(q2_path)
    name, ext = os.path.splitext(q2_base)
    out_path = os.path.join(q2_dir, f"{name} (change highlighted){ext}")

    if progress_cb: progress_cb(f"Saving: {out_path}")
    wb.save(out_path)
    return sheet_name, key_header_q1, key_header_q2, out_path

# ---- GUI ----
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("720x320")

        self.q1_path = tk.StringVar()
        self.q2_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.columns_letters = tk.StringVar(value="I,O,P,Q,R")  # prefill; user can change

        pad = {'padx': 10, 'pady': 6}
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, **pad)

        # First workbook
        ttk.Label(frm, text="First workbook (older):").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.q1_path, width=70).grid(row=0, column=1, sticky="we")
        ttk.Button(frm, text="Browse...", command=self.browse_q1).grid(row=0, column=2, sticky="we")

        # Second workbook
        ttk.Label(frm, text="Second workbook (newer):").grid(row=1, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.q2_path, width=70).grid(row=1, column=1, sticky="we")
        ttk.Button(frm, text="Browse...", command=self.browse_q2).grid(row=1, column=2, sticky="we")

        # Column letters (required)
        ttk.Label(frm, text="Column letters to compare (required):").grid(row=2, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.columns_letters, width=25).grid(row=2, column=1, sticky="w")
        ttk.Label(frm, text='Example: I,O,P,Q,R (include O,S,T for word-level diff)').grid(row=2, column=2, sticky="w")

        # Run
        ttk.Button(frm, text="Run comparison", command=self.run_compare).grid(row=3, column=0, columnspan=3, sticky="we", pady=10)

        # Status
        self.status = tk.StringVar(value="Ready")
        ttk.Label(frm, textvariable=self.status, foreground="gray").grid(row=4, column=0, columnspan=3, sticky="w")

        for i in range(3):
            frm.grid_columnconfigure(i, weight=1)

    def browse_q1(self):
        path = filedialog.askopenfilename(title="Select First Workbook", filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if path:
            self.q1_path.set(path)

    def browse_q2(self):
        path = filedialog.askopenfilename(title="Select Second Workbook", filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if path:
            self.q2_path.set(path)

    def _set_status(self, msg):
        self.status.set(msg)
        self.update_idletasks()

    def run_compare(self):
        q1 = self.q1_path.get().strip()
        q2 = self.q2_path.get().strip()
        sheet = self.sheet_name.get().strip() or None
        cols_raw = self.columns_letters.get().strip()

        if not q1 or not os.path.exists(q1):
            messagebox.showerror(APP_TITLE, "Please choose a valid First workbook.")
            return
        if not q2 or not os.path.exists(q2):
            messagebox.showerror(APP_TITLE, "Please choose a valid Second workbook.")
            return
        if not cols_raw:
            messagebox.showerror(APP_TITLE, "Please provide column letters to compare (e.g., I,O,P,Q,R).")
            return

        letters = tuple([c.strip().upper() for c in cols_raw.split(",") if c.strip()])
        if not letters:
            messagebox.showerror(APP_TITLE, "No valid column letters found (e.g., I,O,P,Q,R).")
            return

        try:
            def progress_cb(msg):
                self._set_status(msg)
            sheet_used, key_q1, key_q2, outp = highlight_q2_changes(
                q1, q2, target_letters=letters, sheet_name=sheet, progress_cb=progress_cb
            )
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Error: {e}")
            self._set_status("Error")
            return

        self._set_status("Done")
        messagebox.showinfo(APP_TITLE, f"Completed!\n\nSheet: {sheet_used}\nSaved: {outp}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
